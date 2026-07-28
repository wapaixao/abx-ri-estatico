import json, re, math
from pathlib import Path
from functools import lru_cache
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries

ROOT = Path('/root/content/sites/abx-ri-estatico')
DATA_PATH = ROOT/'data.json'
DRU_XLSX = Path('/root/data/abx/entregas/APRESENTACAO/ABX_DRU_RI_Receita_Gerencial_U006_VALIDACAO_V3_AUDITADA.xlsx')
U006_XLSX = Path('/root/data/abx/entregas/APRESENTACAO/ABX_Receita_Gerencial_U006_2T2026_VALIDACAO.xlsx')
PISCOFINS_XLSX = Path('/root/data/abx/APURACAO_COFINS_PIS_2T2026_COMPLETA_FORMATADA_SEM_OBS.xlsx')

REF_RE = re.compile(r"(?:(?:'([^']+)'|([A-Za-z0-9_À-ÿ ]+))!)?(\$?[A-Z]{1,3}\$?[0-9]{1,5})(?![A-Za-z0-9_])")
RANGE_RE = re.compile(r"(?:(?:'([^']+)'|([A-Za-z0-9_À-ÿ ]+))!)?(\$?[A-Z]{1,3}\$?[0-9]{1,5}):(\$?[A-Z]{1,3}\$?[0-9]{1,5})")

def split_args(s):
    args=[]; cur=''; depth=0
    for ch in s:
        if ch=='(': depth+=1; cur+=ch
        elif ch==')': depth-=1; cur+=ch
        elif ch==',' and depth==0: args.append(cur); cur=''
        else: cur+=ch
    args.append(cur)
    return args

class EvalBook:
    def __init__(self, path):
        self.wb = openpyxl.load_workbook(path, data_only=False, read_only=False)
    def clean_addr(self, addr): return addr.replace('$','')
    @lru_cache(None)
    def val(self, sheet, addr):
        addr=self.clean_addr(addr)
        v=self.wb[sheet][addr].value
        if v is None: return 0.0
        if isinstance(v,(int,float)): return float(v)
        if isinstance(v,str) and v.startswith('='):
            try: return self.eval_formula(sheet, v[1:])
            except Exception: return 0.0
        return v
    def range_sum(self, sheet, rng):
        min_col,min_row,max_col,max_row=range_boundaries(rng.replace('$',''))
        total=0.0
        for r in range(min_row,max_row+1):
            for c in range(min_col,max_col+1):
                v=self.val(sheet, f'{get_column_letter(c)}{r}')
                if isinstance(v,(int,float)): total += v
        return total
    def eval_expr(self, sheet, expr):
        expr=str(expr).strip()
        # SUM(...)
        while re.search(r'\bSUM\(', expr, flags=re.I):
            m=re.search(r'\bSUM\(', expr, flags=re.I)
            start=m.end(); depth=1; i=start
            while i < len(expr) and depth:
                if expr[i]=='(': depth+=1
                elif expr[i]==')': depth-=1
                i+=1
            inside=expr[start:i-1]
            total=0.0
            for arg in split_args(inside):
                arg=arg.strip()
                rm=RANGE_RE.fullmatch(arg)
                if rm:
                    sh=rm.group(1) or rm.group(2) or sheet
                    total += self.range_sum(sh, f'{rm.group(3)}:{rm.group(4)}')
                else:
                    v=self.eval_expr(sheet,arg)
                    if isinstance(v,(int,float)): total += v
            expr=expr[:m.start()] + str(total) + expr[i:]
        expr=re.sub(r'(\d+(?:\.\d+)?)%', lambda m: str(float(m.group(1))/100), expr)
        def repl_ref(m):
            sh=m.group(1) or m.group(2) or sheet
            v=self.val(sh, m.group(3))
            return str(v if isinstance(v,(int,float)) else 0)
        expr=REF_RE.sub(repl_ref, expr).replace('^','**')
        return float(eval(expr, {'__builtins__':{}}, {}))
    def eval_formula(self, sheet, f):
        f=str(f).strip()
        if f.upper().startswith('IFERROR('):
            args=split_args(f[8:-1])
            try: return self.eval_expr(sheet,args[0])
            except Exception: return self.eval_expr(sheet,args[1]) if len(args)>1 else 0.0
        if f.upper().startswith('IF('):
            args=split_args(f[3:-1])
            cond=REF_RE.sub(lambda m: str(self.val(m.group(1) or m.group(2) or sheet,m.group(3))), args[0].replace('%','/100'))
            ok=eval(cond, {'__builtins__':{}}, {})
            return self.eval_expr(sheet,args[1] if ok else args[2])
        return self.eval_expr(sheet,f)

def clean_num(v):
    if isinstance(v,(int,float)):
        if math.isfinite(v) and abs(v-round(v))<1e-7: return int(round(v))
        return float(v)
    return v

def cell_value(eb, sheet, r, c):
    v=eb.wb[sheet].cell(r,c).value
    if isinstance(v,str) and v.startswith('='):
        formula=v[1:].strip()
        m=REF_RE.fullmatch(formula)
        if m:
            v=eb.val(m.group(1) or m.group(2) or sheet, m.group(3))
        else:
            v=eb.val(sheet, f'{get_column_letter(c)}{r}')
    return clean_num(v)

def extract_dru():
    eb=EvalBook(DRU_XLSX); ws=eb.wb['Trimestral - Valores']
    # value columns identified by header row 6 == Valor
    companies=[]; colmap=[]; current=None
    for c in range(3, ws.max_column+1):
        h=ws.cell(4,c).value
        if h: current=str(h).strip()
        if str(ws.cell(6,c).value or '').strip().upper()=='VALOR':
            period=str(ws.cell(5,c).value or '').strip()
            if current and period:
                if current not in [x['name'] for x in companies]:
                    code=current.split(' - ')[0].replace('TOTAL FILIAIS 001-011','TOTAL').replace('TOTAL GRUPO','TOTAL')
                    companies.append({'name':current,'code':code})
                colmap.append((current, period, c))
    periods=[]
    for _,p,_ in colmap:
        if p not in periods: periods.append(p)
    rows=[]
    for r in range(7, ws.max_row+1):
        if ws.row_dimensions[r].hidden: continue
        desc=str(ws.cell(r,2).value or '').strip()
        code=str(ws.cell(r,1).value or '').strip()
        if not desc and not code: continue
        dupper=desc.upper()
        if dupper in ['DISTRIBUICAO LUCRO','DISTRIBUICAO DE LUCROS','PARTICIPACAO NOS RESULTADOS']: continue
        empresas={c['name']:{p:0 for p in periods} for c in companies}
        for comp,p,c in colmap:
            empresas[comp][p]=cell_value(eb,'Trimestral - Valores',r,c) or 0
        if dupper in ['RECEITA LIQUIDA','RECEITA BRUTA','LUCRO BRUTO','EBITDA','LAIR','LUCRO LIQUIDO','LAIR GERENCIAL','NOVO LUCRO LÍQUIDO','LUCRO LÍQUIDO GERENCIAL']:
            nivel=1
        elif len(code)<=4 or code.startswith('AJ') or not code:
            nivel=2
        else:
            nivel=3
        rows.append({'codigo':code,'secao':'DRE/DRU','nivel':nivel,'descricao':desc,'empresas':empresas,'grupo':empresas.get('TOTAL GRUPO',{})})
    return {'label':'Demonstração do Resultado da Unidade (DRU)','periods':periods,'companies':companies,'rows':rows}

def fill_hex(cell):
    fg=cell.fill.fgColor
    if fg.type=='rgb' and fg.rgb and fg.rgb!='00000000': return '#'+fg.rgb[-6:]
    return ''

def font_hex(cell):
    fc=cell.font.color
    try:
        if fc and fc.type=='rgb' and fc.rgb: return '#'+fc.rgb[-6:]
    except Exception: pass
    return ''

def extract_sheet_report(path, sheet, label, max_row=None, max_col=None):
    eb=EvalBook(path); ws=eb.wb[sheet]
    max_row=max_row or ws.max_row; max_col=max_col or ws.max_column
    merged={}
    covered=set()
    for rng in ws.merged_cells.ranges:
        minc,minr,maxc,maxr=rng.bounds
        if minr>max_row or minc>max_col: continue
        merged[(minr,minc)]=(min(maxc,max_col)-minc+1, min(maxr,max_row)-minr+1)
        for rr in range(minr,min(maxr,max_row)+1):
            for cc in range(minc,min(maxc,max_col)+1):
                if (rr,cc)!=(minr,minc): covered.add((rr,cc))
    rows=[]
    for r in range(1,max_row+1):
        if ws.row_dimensions[r].hidden: continue
        row=[]; any_value=False
        for c in range(1,max_col+1):
            if (r,c) in covered: continue
            cell=ws.cell(r,c)
            v=cell_value(eb,sheet,r,c)
            if v not in [None,'']: any_value=True
            cs,rs=merged.get((r,c),(1,1))
            style={}
            bg=fill_hex(cell); color=font_hex(cell)
            if bg: style['bg']=bg
            if color: style['color']=color
            if cell.font.bold: style['bold']=True
            row.append({'v':v if v is not None else '', 'cs':cs, 'rs':rs, 'style':style})
        if any_value or r<=5:
            rows.append(row)
    return {'label':label,'type':'sheet','rows':rows}

def extract_piscofins_control():
    eb=EvalBook(PISCOFINS_XLSX); ws=eb.wb['Cofins e Pis']
    block_rows=[2,11,20,29]
    periods=[]; units=[]; rows_by_desc={}; desc_order=[]
    for br in block_rows:
        period=str(ws.cell(br,1).value or '').strip()
        if not period: continue
        periods.append(period)
        local_units=[]
        for c in range(2, ws.max_column+1):
            name=str(ws.cell(br,c).value or '').strip()
            if not name: continue
            local_units.append((name,c))
            if name not in units: units.append(name)
        for rr in range(br+1, br+6):
            desc=str(ws.cell(rr,1).value or '').strip()
            if not desc: continue
            if desc not in rows_by_desc:
                rows_by_desc[desc]={}; desc_order.append(desc)
            for u in units:
                rows_by_desc[desc].setdefault(u,{})
                for p in periods: rows_by_desc[desc][u].setdefault(p,0)
            for name,c in local_units:
                rows_by_desc[desc][name][period]=cell_value(eb,'Cofins e Pis',rr,c) or 0
    companies=[{'name':u,'code':u.split(' - ')[0]} for u in units]
    outrows=[]
    for desc in desc_order:
        nivel=1 if desc.upper()=='TOTAL' else 2
        outrows.append({'descricao':desc,'nivel':nivel,'empresas':rows_by_desc[desc]})
    return {'label':'PIS / COFINS','type':'unit_period_control','periods':periods,'companies':companies,'rows':outrows}

def apply_2026_negative_previous_rule(report):
    # For Distribuição 2026 blocks only, move negative Resultado to Negativo Anterior.
    for ridx,row in enumerate(report['rows']):
        for i,cell in enumerate(row):
            v=cell.get('v')
            if isinstance(v,str) and '2026' in v:
                try:
                    resultado=report['rows'][ridx+1][i+1]['v']
                    neg_cell=report['rows'][ridx+2][i+1]
                    liq_cell=report['rows'][ridx+3][i+1]
                except Exception:
                    continue
                if isinstance(resultado,(int,float)) and resultado < 0:
                    neg_cell['v']=resultado
                    liq_cell['v']=resultado
    return report

def unit_code_from_label(label):
    s=str(label or '')
    m=re.search(r'(\d{3})', s)
    return m.group(1) if m else None

def apply_dru_lucro_to_distrib(report, dru_report):
    """For 2026 distribution blocks, use DRU Lucro Líquido as Resultado/Lucro source."""
    lucro_rows=[r for r in dru_report['rows'] if str(r.get('descricao','')).strip().upper()=='LUCRO LIQUIDO']
    if not lucro_rows: return report
    lucro=lucro_rows[0]['empresas']
    dru_by_code={unit_code_from_label(name): vals for name, vals in lucro.items() if unit_code_from_label(name) and not str(name).upper().startswith('TOTAL')}
    header_units=[c.get('v') for c in report['rows'][0] if unit_code_from_label(c.get('v'))]
    periods={'1º TRIM 2026':'1T26','2º TRIM 2026':'2T26'}
    for ridx,row in enumerate(report['rows']):
        vals=[c.get('v') for c in row]
        period_label=next((v for v in vals if isinstance(v,str) and v in periods), None)
        if not period_label: continue
        period=periods[period_label]
        result_label_positions=[i for i,c in enumerate(report['rows'][ridx+1]) if str(c.get('v')).strip().upper()=='RESULTADO']
        partner_label_positions=[i for i,c in enumerate(report['rows'][ridx+4]) if str(c.get('v')).strip().upper()=='PARCEIRO'] if ridx+4 < len(report['rows']) else []
        for idx,label_pos in enumerate(result_label_positions):
            if idx >= len(header_units): break
            code=unit_code_from_label(header_units[idx])
            val=float(dru_by_code.get(code,{}).get(period,0) or 0)
            value_pos=label_pos+1
            if value_pos < len(report['rows'][ridx+1]):
                report['rows'][ridx+1][value_pos]['v']=val if val>0 else 0
            if value_pos < len(report['rows'][ridx+2]):
                report['rows'][ridx+2][value_pos]['v']=val if val<0 else 0
            if value_pos < len(report['rows'][ridx+3]):
                report['rows'][ridx+3][value_pos]['v']=val
            # Distribution values: only distribute positive Resultado Líquido.
            if idx < len(partner_label_positions):
                ppos=partner_label_positions[idx]
                for pr in range(ridx+5, min(ridx+10, len(report['rows']))):
                    if ppos+2 >= len(report['rows'][pr]): continue
                    pct=report['rows'][pr][ppos+1].get('v')
                    partner=report['rows'][pr][ppos].get('v')
                    if partner in ['', None] or not isinstance(pct,(int,float)):
                        if report['rows'][pr][ppos+2].get('v') not in ['', None]: report['rows'][pr][ppos+2]['v']=''
                        continue
                    report['rows'][pr][ppos+2]['v']=(val*pct/100) if val>0 else 0
    return report

def main():
    data=json.loads(DATA_PATH.read_text())
    reports=data.setdefault('reports',{})
    reports['DRU']=extract_dru()
    reports['U006']=extract_sheet_report(U006_XLSX,'Receita Gerencial U006','Receita Gerencial U006',max_row=11,max_col=76)
    reports['PISCOFINS']=extract_piscofins_control()
    reports['DISTRIB']=apply_dru_lucro_to_distrib(extract_sheet_report(PISCOFINS_XLSX,'APRESENTAÇÃO','Distribuição de Resultado',max_row=51,max_col=60), reports['DRU'])
    DATA_PATH.write_text(json.dumps(data,ensure_ascii=False,indent=2),encoding='utf-8')
    print('wrote',DATA_PATH)
    print('reports',list(reports.keys()))
    print('DRU rows',len(reports['DRU']['rows']),'PIS rows',len(reports['PISCOFINS']['rows']),'DISTR rows',len(reports['DISTRIB']['rows']))

if __name__=='__main__': main()
